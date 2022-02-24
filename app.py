from reportlab.pdfgen import canvas
import pandas as pd
import os
import openpyxl
import msvcrt

#Function declaration
def msg_init():
    msg = '''\n Hola, para que el programa funcione correctamente y genere todos los carteles debemos tener en cuenta que: \n
    1. Las plantillas OBSERVACIONES PEDIDOS.xlsx y CL - Autocontrol Producción DV - Ed. 02.xlsx esten en el mismo direcorio que el .exe\n	
    2. No tener abierto ningun cartel, protocolo u hoja de observaciones abierta.
    \n \n >>>>> Presione una tecla para continuar...'''
    print(msg)
    msvcrt.getch()
def msg_error(co, line, tipo):
    print("======================================================================")
    print("|>>> ERROR AL CREAR ==> " + co[line] + "_" + tipo + " <<<|")
    print("======================================================================")

#Sheets sizes#widht
width = 841
height = 595

width2 = 595
height2 = 841

A4 = [width, height]
A4Folder = [width2, height2]

#Destination folder of PDF and Excel files created
pathDestination = "U:/OPERACIONES/08 FÁBRICA/1 AUTOMATIZACIÓN CARTELES TRABAJO JEFES TURNO/"

msg_init()

#If destination folder does not exist, the folder is created automatically
if not os.path.exists(pathDestination):
    os.makedirs(pathDestination)

#The info input is trough of excel file
excel = "U:/OPERACIONES/05 PLANIFICACIÓN/SEGUIMIENTO_PEDIDOS_V04.xlsm"
df = pd.read_excel(excel, sheet_name= "AHU", skiprows=1)

#Excel template
excel_observaciones = openpyxl.load_workbook("OBSERVACIONES PEDIDOS.xlsx")
sheet1_observaciones = excel_observaciones.active

#Excel template
excel_protocolo = openpyxl.load_workbook("CL - Autocontrol Producción DV - Ed. 02.xlsx")
sheet1_protocolo = excel_protocolo.active

#Posters titles
title = "PANELES Y PUERTAS"
title2 = "PERFILES"
title3 = "CARPETA"

#In "SEGUIMIENTO_PEDIDOS.xlsm" file we get in variables the differents columns values we need.
co = df["Unit"].values
mo = df["MO no"].values
model = df["CO Item no"].values
status = df["MO sts"].values
date = df["MO Start"].values
country = df["Country"].values

#We create a dictionary "ISO2 : Country"
isoCountry = {
    "AF":"Afganistán",
    "AL":"Albania",
    "DE":"Alemania",
    "DZ":"Algeria",
    "AD":"Andorra",
    "AO":"Angola",
    "AI":"Anguila",
    "AQ":"Antártida",
    "AG":"Antigua y Barbuda",
    "AN":"Antillas Neerlandesas",
    "SA":"Arabia Saudita",
    "AR":"Argentina",
    "AM":"Armenia",
    "AW":"Aruba",
    "AU":"Australia",
    "AT":"Austria",
    "AZ":"Azerbayán",
    "BE":"Bélgica",
    "BS":"Bahamas",
    "BH":"Bahrein",
    "BD":"Bangladesh",
    "BB":"Barbados",
    "BZ":"Belice",
    "BJ":"Benín",
    "BT":"Bhután",
    "BY":"Bielorrusia",
    "MM":"Birmania",
    "BO":"Bolivia",
    "BA":"Bosnia y Herzegovina",
    "BW":"Botsuana",
    "BR":"Brasil",
    "BN":"Brunéi",
    "BG":"Bulgaria",
    "BF":"Burkina Faso",
    "BI":"Burundi",
    "CV":"Cabo Verde",
    "KH":"Camboya",
    "CM":"Camerún",
    "CA":"Canadá",
    "TD":"Chad",
    "CL":"Chile",
    "CN":"China",
    "CY":"Chipre",
    "VA":"Ciudad del Vaticano",
    "CO":"Colombia",
    "KM":"Comoras",
    "CG":"Congo",
    "CD":"Congo",
    "KP":"Corea del Norte",
    "KR":"Corea del Sur",
    "CI":"Costa de Marfil",
    "CR":"Costa Rica",
    "HR":"Croacia",
    "CU":"Cuba",
    "DK":"Dinamarca",
    "DM":"Dominica",
    "EC":"Ecuador",
    "EG":"Egipto",
    "SV":"El Salvador",
    "AE":"Emiratos Árabes Unidos",
    "ER":"Eritrea",
    "SK":"Eslovaquia",
    "SI":"Eslovenia",
    "ES":"España",
    "US":"Estados Unidos de América",
    "EE":"Estonia",
    "ET":"Etiopía",
    "PH":"Filipinas",
    "FI":"Finlandia",
    "FJ":"Fiyi",
    "FR":"Francia",
    "GA":"Gabón",
    "GM":"Gambia",
    "GE":"Georgia",
    "GH":"Ghana",
    "GI":"Gibraltar",
    "GD":"Granada",
    "GR":"Grecia",
    "GL":"Groenlandia",
    "GP":"Guadalupe",
    "GU":"Guam",
    "GT":"Guatemala",
    "GF":"Guayana Francesa",
    "GG":"Guernsey",
    "GN":"Guinea",
    "GQ":"Guinea Ecuatorial",
    "GW":"Guinea-Bissau",
    "GY":"Guyana",
    "HT":"Haití",
    "HN":"Honduras",
    "HK":"Hong kong",
    "HU":"Hungría",
    "IN":"India",
    "ID":"Indonesia",
    "IR":"Irán",
    "IQ":"Irak",
    "IE":"Irlanda",
    "BV":"Isla Bouvet",
    "IM":"Isla de Man",
    "CX":"Isla de Navidad",
    "NF":"Isla Norfolk",
    "IS":"Islandia",
    "BM":"Islas Bermudas",
    "KY":"Islas Caimán",
    "CC":"Islas Cocos (Keeling)",
    "CK":"Islas Cook",
    "AX":"Islas de Åland",
    "FO":"Islas Feroe",
    "GS":"Islas Georgias del Sur y Sandwich del Sur",
    "HM":"Islas Heard y McDonald",
    "MV":"Islas Maldivas",
    "FK":"Islas Malvinas",
    "MP":"Islas Marianas del Norte",
    "MH":"Islas Marshall",
    "PN":"Islas Pitcairn",
    "SB":"Islas Salomón",
    "TC":"Islas Turcas y Caicos",
    "UM":"Islas Ultramarinas Menores de Estados Unidos",
    "VG":"Islas Vírgenes Británicas",
    "VI":"Islas Vírgenes de los Estados Unidos",
    "IL":"Israel",
    "IT":"Italia",
    "JM":"Jamaica",
    "JP":"Japón",
    "JE":"Jersey",
    "JO":"Jordania",
    "KZ":"Kazajistán",
    "KE":"Kenia",
    "KG":"Kirgizstán",
    "KI":"Kiribati",
    "KW":"Kuwait",
    "LB":"Líbano",
    "LA":"Laos",
    "LS":"Lesoto",
    "LV":"Letonia",
    "LR":"Liberia",
    "LY":"Libia",
    "LI":"Liechtenstein",
    "LT":"Lituania",
    "LU":"Luxemburgo",
    "MX":"México",
    "MC":"Mónaco",
    "MO":"Macao",
    "MK":"Macedônia",
    "MG":"Madagascar",
    "MY":"Malasia",
    "MW":"Malawi",
    "ML":"Mali",
    "MT":"Malta",
    "MA":"Marruecos",
    "MQ":"Martinica",
    "MU":"Mauricio",
    "MR":"Mauritania",
    "YT":"Mayotte",
    "FM":"Micronesia",
    "MD":"Moldavia",
    "MN":"Mongolia",
    "ME":"Montenegro",
    "MS":"Montserrat",
    "MZ":"Mozambique",
    "NA":"Namibia",
    "NR":"Nauru",
    "NP":"Nepal",
    "NI":"Nicaragua",
    "NE":"Niger",
    "NG":"Nigeria",
    "NU":"Niue",
    "NO":"Noruega",
    "NC":"Nueva Caledonia",
    "NZ":"Nueva Zelanda",
    "OM":"Omán",
    "NL":"Países Bajos",
    "PK":"Pakistán",
    "PW":"Palau",
    "PS":"Palestina",
    "PA":"Panamá",
    "PG":"Papúa Nueva Guinea",
    "PY":"Paraguay",
    "PE":"Perú",
    "PF":"Polinesia Francesa",
    "PL":"Polonia",
    "PT":"Portugal",
    "PR":"Puerto Rico",
    "QA":"Qatar",
    "GB":"Reino Unido",
    "CF":"República Centroafricana",
    "CZ":"República Checa",
    "DO":"República Dominicana",
    "RE":"Reunión",
    "RW":"Ruanda",
    "RO":"Rumanía",
    "RU":"Rusia",
    "EH":"Sahara Occidental",
    "WS":"Samoa",
    "AS":"Samoa Americana",
    "BL":"San Bartolomé",
    "KN":"San Cristóbal y Nieves",
    "SM":"San Marino",
    "MF":"San Martín (Francia)",
    "PM":"San Pedro y Miquelón",
    "VC":"San Vicente y las Granadinas",
    "SH":"Santa Elena",
    "LC":"Santa Lucía",
    "ST":"Santo Tomé y Príncipe",
    "SN":"Senegal",
    "RS":"Serbia",
    "SC":"Seychelles",
    "SL":"Sierra Leona",
    "SG":"Singapur",
    "SY":"Siria",
    "SO":"Somalia",
    "LK":"Sri lanka",
    "ZA":"Sudáfrica",
    "SD":"Sudán",
    "SE":"Suecia",
    "CH":"Suiza",
    "SR":"Surinám",
    "SJ":"Svalbard y Jan Mayen",
    "SZ":"Swazilandia",
    "TJ":"Tadjikistán",
    "TH":"Tailandia",
    "TW":"Taiwán",
    "TZ":"Tanzania",
    "IO":"Territorio Británico del Océano Índico",
    "TF":"Territorios Australes y Antárticas Franceses",
    "TL":"Timor Oriental",
    "TG":"Togo",
    "TK":"Tokelau",
    "TO":"Tonga",
    "TT":"Trinidad y Tobago",
    "TN":"Tunez",
    "TM":"Turkmenistán",
    "TR":"Turquía",
    "TV":"Tuvalu",
    "UA":"Ucrania",
    "UG":"Uganda",
    "UY":"Uruguay",
    "UZ":"Uzbekistán",
    "VU":"Vanuatu",
    "VE":"Venezuela",
    "VN":"Vietnam",
    "WF":"Wallis y Futuna",
    "YE":"Yemen",
    "DJ":"Yibuti",
    "ZM":"Zambia",
}

#We read every line in "SEGUIMIENTO DE PEDIDOS" file. 
for line in range(len(mo)):
    #Only need to read the status different to "90-90" lines.
    if status[line] != "90-90":
        moFloat = f"MO: {mo[line]:.0f}"
        #Creation of the poster of doors and pannels.
        try:
            pdf_Puertas = canvas.Canvas(pathDestination + co[line] + "_" + "PUERTAS" + ".pdf", pagesize=A4)
            pdf_Puertas.setFont('Helvetica-Bold', 72)
            pdf_Puertas.drawCentredString(width/2, height - 100, title)
            pdf_Puertas.drawCentredString(width/2, height - 220, "PEDIDO: " + co[line])
            pdf_Puertas.drawCentredString(width/2, height - 340, moFloat)
            pdf_Puertas.drawCentredString(width/2, height - 460, model[line])
            pdf_Puertas.setFont('Helvetica-Bold', 60)
            pdf_Puertas.drawCentredString(width/2, height - 580, isoCountry[country[line]])
            pdf_Puertas.save()
        except:
            msg_error(co, line, "PUERTAS.pdf")

        #Creation of the poster of profiles.
        try:
            pdf_Perfiles = canvas.Canvas(
            pathDestination + co[line] + "_" + title2 + ".pdf", pagesize=A4)
            pdf_Perfiles.setFont('Helvetica-Bold', 72)
            pdf_Perfiles.drawCentredString(width/2, height - 100, title2)
            pdf_Perfiles.drawCentredString(width/2, height - 220, "PEDIDO: " + co[line])
            pdf_Perfiles.drawCentredString(width/2, height - 340, moFloat)
            pdf_Perfiles.drawCentredString(width/2, height - 460, model[line])
            pdf_Perfiles.setFont('Helvetica-Bold', 60)
            pdf_Perfiles.drawCentredString(width/2, height - 580, isoCountry[country[line]])
            pdf_Perfiles.save()
        except:
            msg_error(co, line, "PERFILES.pdf")

        #Creation of the poster of folder
        try:
            pdf_Folder = canvas.Canvas(
            pathDestination + co[line] + "_" + title3 + ".pdf", pagesize=A4Folder)
            pdf_Folder.setFont('Helvetica-Bold', 50)
            pdf_Folder.drawCentredString(width2/2, height2-60, "CO: " + co[line])
            pdf_Folder.drawCentredString(width2/2, height2-150, model[line])
            pdf_Folder.drawCentredString(width2/2, height2-270, moFloat)
            pdf_Folder.drawCentredString(width2/2, height2-390, isoCountry[country[line]])
            pdf_Folder.save()
        except:
            msg_error(co, line, "CARPETA.pdf")

        #Creation of Excel and cell fill.
        try:
            celdaA4_observaciones = sheet1_observaciones.cell(row=4, column=1)
            celdaA4_observaciones.value = co[line]
            celdaC4_observaciones = sheet1_observaciones.cell(row=4, column=3)
            celdaC4_observaciones.value = model[line]
            celdaD4_observaciones = sheet1_observaciones.cell(row=4, column=4)
            celdaD4_observaciones.value = mo[line]
            celdaE4_observaciones = sheet1_observaciones.cell(row= 4, column=5)
            celdaE4_observaciones.value = isoCountry[country[line]]
            excel_observaciones.save(pathDestination + co[line] + "_OBSERVACIONES.xlsx")
        except:
            msg_error(co, line, "OBSERVACIONES.xlsx")
            
        #Creation of Excel and cell fill.
        try:
            celdaB4_protocolo = sheet1_protocolo.cell(row=4, column=2)
            celdaB4_protocolo.value = co[line]
            celdaC4_protocolo = sheet1_protocolo.cell(row= 4, column=4)
            celdaC4_protocolo.value = moFloat
            celdaB5_protocolo = sheet1_protocolo.cell(row=5, column=2)
            celdaB5_protocolo.value = model[line]
            excel_protocolo.save(pathDestination + co[line] + "_PROTOCOLO.xlsx")
        except:
            msg_error(co, line, "PROTOCOLO.xlsx")

        #Print on console the CO, MO, model and country.
        print(co[line] + " >> " + moFloat + " >> " + model[line] + " >> " + isoCountry[country[line]])
